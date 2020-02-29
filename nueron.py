from openpyxl import Workbook
import random
import xlrd

def trainingSet():
    # TODO Creating data set
    #     1. Data set will be a list of numbers either 0 or 1:
    #     Example 1:
    #           input: 0 1 1       output: 0
    #           input: 1 0 0       output: 1
    #           input: 0 0 0       output: 1

    # TODO Creating the nueral node class
    #    2. Write a nueral class object
    #       - Weights increase: Wx + b > 0 -> 1
    #                           Wx + b < 0 -> 0
    #       - Start with the same wieght of type double either from 0-1 maybe 0.5
    #       -

    # TODO Write a method to print out letters to excel sheet to create
    # training set data in which we give it the expected output
    # written in two programs one will be the student and the other will
    # be the teacher.

    book = Workbook()
    trainingSheet = book.active

    count = 0
    val = input("Enter your a number for input values: ")

    if val == "q":
        exit()

    print("printing letters... " + val)

    for i in range(int(val)):
        count = count + 1
        trainingSheet['A' + str(count)] = "G"
        trainingSheet['B' + str(count)] = 'Y'

    book.save("trainingSet.xlsx")
    exit()


def testingSet():
    book = Workbook()
    testingSheet = book.active
    count = 0

    val = input("Enter a number to generate random letters: ")

    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    if val == "q":
        exit()

    for i in range(int(val)):
        count = count + 1
        ramdomLetter = str(random.choice(letters))
        testingSheet['B' + str(count)] = 'N'
        if ramdomLetter == 'G':
            testingSheet['B' + str(count)] = 'Y'
        testingSheet['A' + str(count)] = ramdomLetter

    book.save("testingSet.xlsx")
    exit()

def expectedResults():
    # TODO this is where we will create an expected output for training set
    # TODO also create another training set to give random letters but also if its a G give output of yes
    book = Workbook()

def node():
    val = input("Enter the name of the worksheet to give node: ")

    if val == 'q':
        exit()

    book = xlrd.open_workbook(val)
    sheet = book.sheet_by_name('Sheet')

    dictionary = {
        'A': '0.5',
        'B': '0.5',
        'C': '0.5',
        'D': '0.5',
        'E': '0.5',
        'F': '0.5',
        'G': '0.5',
        'H': '0.5',
        'I': '0.5',
        'J': '0.5',
        'K': '0.5',
        'L': '0.5',
        'M': '0.5',
        'N': '0.5',
        'O': '0.5',
        'P': '0.5',
        'Q': '0.5',
        'R': '0.5',
        'S': '0.5',
        'T': '0.5',
        'U': '0.5',
        'V': '0.5',
        'W': '0.5',
        'X': '0.5',
        'Y': '0.5',
        'Z': '0.5'
    }

    for cell in sheet.col(0):
        letter = cell.value
        dictionary[letter] = str(float(dictionary[letter]) + 0.1)


    exit()

def teacher():
    workbook = xlrd.open_workbook("trainingSet.xlsx")

    sheet = workbook.sheet_by_name('Sheet')

    # for cell in sheet.col(1):
    #     print(cell.value)